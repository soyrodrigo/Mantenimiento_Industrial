#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bot de Mantenimiento Industrial para Telegram
Versión mejorada con soporte para fotos y mejor manejo de errores
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Optional
import uuid

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, ApplicationBuilder, MessageHandler, filters, ContextTypes,
    CommandHandler, ConversationHandler, CallbackQueryHandler
)
import re
import threading
from flask import Flask, jsonify, request
from flask_cors import CORS
import signal

def escape_markdown(text):
    """Escapa caracteres especiales para MarkdownV2"""
    # Caracteres que necesitan escape en MarkdownV2
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    
    # Escapar cada carácter especial
    for char in escape_chars:
        text = text.replace(char, f'\\{char}')
    
    return text

# Función alternativa más simple - usar HTML en lugar de Markdown
def format_html(text):
    """Convierte formato básico a HTML"""
    # Reemplazar **texto** por <b>texto</b>
    text = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', text)
    # Reemplazar *texto* por <i>texto</i>
    text = re.sub(r'\*(.*?)\*', r'<i>\1</i>', text)
    
    return text

# --- CONFIGURACIÓN ---
TOKEN = "7769871938:AAH8NxORtMoC3Q0Rf8n4Ctl9Y056octRTSk"  # Reemplaza con tu token real del bot
ARCHIVO_EXCEL = "registro_mantenimiento.xlsx"
ARCHIVO_CHECKLISTS = "checklists_maquinas.json"
ARCHIVO_CONFIG = "config.json"
CARPETA_FOTOS = "fotos_mantenimiento"  # Nueva carpeta para fotos
ADMIN_IDS = [1214237764]  # Lista de IDs de administradores



# Estados para conversaciones
(ESPERANDO_NOMBRE_MAQUINA, ESPERANDO_ITEMS_CHECKLIST, 
 ESPERANDO_OBSERVACIONES, ESPERANDO_NUEVO_ADMIN, 
 ESPERANDO_FOTO) = range(5)  # Agregado estado para fotos

# Configuración de logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Almacenamiento en memoria
checklists_activos: Dict = {}

# VARIABLES GLOBALES PARA LA API
bot_stats = {
    "checklists_completados": 0,
    "fotos_guardadas": 0,
    "maquinas_configuradas": 0,
    "ultimo_checklist": None,
    "inicio_bot": None
}
# Crear carpeta de fotos si no existe
if not os.path.exists(CARPETA_FOTOS):
    os.makedirs(CARPETA_FOTOS)
    logger.info(f"Creada carpeta de fotos: {CARPETA_FOTOS}")

class ConfigManager:
    """Maneja la configuración del bot"""
    
    @staticmethod
    def cargar_config():
        """Carga configuración desde archivo JSON"""
        if os.path.exists(ARCHIVO_CONFIG):
            try:
                with open(ARCHIVO_CONFIG, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                logger.error(f"Error cargando configuración: {e}")
        return {"admin_ids": ADMIN_IDS}
    
    @staticmethod
    def guardar_config(config):
        """Guarda configuración en archivo JSON"""
        try:
            with open(ARCHIVO_CONFIG, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except IOError as e:
            logger.error(f"Error guardando configuración: {e}")

class FotoManager:
    """Maneja el almacenamiento y organización de fotos"""
    
    @staticmethod
    def generar_nombre_foto(maquina, item, user_id):
        """Genera un nombre único para la foto"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Limpiar nombre de máquina para usar en archivo
        maquina_limpia = "".join(c for c in maquina if c.isalnum() or c in (' ', '-', '_')).rstrip()
        maquina_limpia = maquina_limpia.replace(' ', '_')
        
        # Crear ID único
        unique_id = str(uuid.uuid4())[:8]
        
        return f"{timestamp}_{maquina_limpia}_{user_id}_{unique_id}.jpg"
    
    @staticmethod
    async def guardar_foto(file, nombre_archivo):
        """Guarda la foto en el sistema de archivos"""
        try:
            ruta_completa = os.path.join(CARPETA_FOTOS, nombre_archivo)
            await file.download_to_drive(ruta_completa)
            logger.info(f"Foto guardada: {ruta_completa}")
            return ruta_completa
        except Exception as e:
            logger.error(f"Error guardando foto: {e}")
            return None
    
    @staticmethod
    def obtener_ruta_relativa(ruta_completa):
        """Obtiene la ruta relativa para almacenar en Excel"""
        if ruta_completa:
            return os.path.relpath(ruta_completa)
        return ""

class ChecklistManager:
    """Maneja los checklists de las máquinas"""
    
    @staticmethod
    def cargar_checklists():
        """Carga checklists desde archivo JSON"""
        if os.path.exists(ARCHIVO_CHECKLISTS):
            try:
                with open(ARCHIVO_CHECKLISTS, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                logger.error(f"Error cargando checklists: {e}")
        return {}
    
    @staticmethod
    def guardar_checklists(checklists):
        """Guarda checklists en archivo JSON"""
        try:
            with open(ARCHIVO_CHECKLISTS, 'w', encoding='utf-8') as f:
                json.dump(checklists, f, ensure_ascii=False, indent=2)
        except IOError as e:
            logger.error(f"Error guardando checklists: {e}")
    
    @staticmethod
    def eliminar_maquina(nombre_maquina):
        """Elimina una máquina del checklist"""
        checklists = ChecklistManager.cargar_checklists()
        if nombre_maquina in checklists:
            del checklists[nombre_maquina]
            ChecklistManager.guardar_checklists(checklists)
            return True
        return False

class ExcelManager:
    """Maneja la creación y escritura del archivo Excel"""
    
    @staticmethod
    def crear_excel_mantenimiento():
        """Crea un nuevo archivo Excel con formato"""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Mantenimiento Diario"
        
        # Título principal
        hoja.merge_cells('A1:J1')  # Ampliado para nueva columna
        hoja['A1'] = f"REGISTRO DE MANTENIMIENTO - {datetime.now().strftime('%Y')}"
        hoja['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        hoja['A1'].fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        hoja['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        hoja.merge_cells('A2:J2')  # Ampliado para nueva columna
        hoja['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        hoja['A2'].font = Font(size=10, italic=True)
        
        # Encabezados - Agregada columna FOTO
        encabezados = ["FECHA", "HORA", "OPERADOR", "MÁQUINA", "ITEM", 
                      "ESTADO", "OBSERVACIONES", "FOTO", "RESULTADO", "DURACIÓN"]
        
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=4, column=col)
            celda.value = encabezado
            celda.font = Font(bold=True, color='FFFFFF')
            celda.fill = PatternFill(start_color='D67228', end_color='D67228', fill_type='solid')
            celda.alignment = Alignment(horizontal='center')
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            celda.border = thin_border
        
        # Ajustar ancho de columnas - Agregado ancho para columna FOTO
        anchos = [12, 10, 15, 20, 30, 12, 25, 20, 15, 12]
        for i, ancho in enumerate(anchos, 1):
            hoja.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        
        hoja.freeze_panes = 'A5'
        return libro, hoja
    
    @staticmethod
    def guardar_resultado_checklist(fecha, hora, operador, maquina, item, 
                                   estado, observaciones, foto_ruta, resultado_final, duracion=None):
        """Guarda un resultado en el archivo Excel - Agregado parámetro foto_ruta"""
        try:
            if not os.path.exists(ARCHIVO_EXCEL):
                libro, hoja = ExcelManager.crear_excel_mantenimiento()
                fila = 5
            else:
                libro = openpyxl.load_workbook(ARCHIVO_EXCEL)
                hoja = libro.active
                fila = hoja.max_row + 1
            
            # Agregada foto_ruta en los datos
            datos = [fecha, hora, operador, maquina, item, estado, 
                    observaciones, foto_ruta, resultado_final, duracion or ""]
            
            for col, dato in enumerate(datos, 1):
                celda = hoja.cell(row=fila, column=col, value=dato)
                # Aplicar color según el estado (columna 6)
                if col == 6:  # Columna de estado
                    if "OK" in str(dato):
                        celda.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif "REVISAR" in str(dato):
                        celda.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    elif "FALLA" in str(dato):
                        celda.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                # Resaltar celdas con foto (columna 8)
                elif col == 8 and dato:
                    celda.fill = PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid')
                    celda.font = Font(color='1976D2')
            
            libro.save(ARCHIVO_EXCEL)
            logger.info(f"Resultado guardado: {maquina} - {item} - Foto: {'Sí' if foto_ruta else 'No'}")
            
        except Exception as e:
            logger.error(f"Error guardando en Excel: {e}")

def es_admin(user_id):
    """Verifica si el usuario es administrador"""
    config = ConfigManager.cargar_config()
    return user_id in config.get("admin_ids", ADMIN_IDS)

async def comando_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando de inicio"""
    user_id = update.message.from_user.id
    nombre = update.message.from_user.first_name
    
    if es_admin(user_id):
        mensaje = f"🔧 Hola {nombre}, eres administrador.\n\n" \
                 "📋 **Comandos disponibles:**\n" \
                 "• /nueva_maquina - Agregar nueva máquina\n" \
                 "• /ver_maquinas - Ver máquinas configuradas\n" \
                 "• /eliminar_maquina - Eliminar máquina\n" \
                 "• /checklist - Realizar checklist\n" \
                 "• /descargar_excel - Descargar registro\n" \
                 "• /estadisticas - Ver estadísticas\n" \
                 "• /agregar_admin - Agregar administrador\n" \
                 "• /ver_fotos - Ver fotos recientes\n" \
                 "• /mi_id - Ver tu ID"
    else:
        mensaje = f"👋 Hola {nombre}!\n\n" \
                 "📋 **Comandos disponibles:**\n" \
                 "• /checklist - Realizar checklist de mantenimiento\n" \
                 "• /mi_id - Ver tu ID de usuario\n\n" \
                 "📸 **Nuevo:** Ahora puedes enviar fotos durante el checklist\n" \
                 "para documentar problemas encontrados."
    
    await update.message.reply_text(mensaje, parse_mode='HTML')

async def comando_mi_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra el ID del usuario"""
    user_id = update.message.from_user.id
    username = update.message.from_user.username or "Sin username"
    await update.message.reply_text(
        f"👤 **Tu información:**\n"
        f"• ID: `{user_id}`\n"
        f"• Username: @{username}",
        parse_mode='HTML'
    )

async def comando_nueva_maquina(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inicia el proceso de agregar una nueva máquina"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden agregar máquinas.")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "🏭 **Agregar nueva máquina**\n\n"
        "Escribe el nombre de la nueva máquina:",
        parse_mode='HTML'
    )
    return ESPERANDO_NOMBRE_MAQUINA

async def recibir_nombre_maquina(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recibe el nombre de la máquina"""
    nombre_maquina = update.message.text.strip()
    
    # Validar que no exista ya
    checklists = ChecklistManager.cargar_checklists()
    if nombre_maquina in checklists:
        await update.message.reply_text(
            f"⚠️ La máquina '{nombre_maquina}' ya existe.\n"
            "Escribe otro nombre:"
        )
        return ESPERANDO_NOMBRE_MAQUINA
    
    context.user_data['nombre_maquina'] = nombre_maquina
    await update.message.reply_text(
        f"✅ Máquina: **{nombre_maquina}**\n\n"
        "📝 Ahora escribe los ítems del checklist, **uno por línea**:\n\n"
        "*Ejemplo:*\n"
        "• Verificar nivel de aceite\n"
        "• Revisar temperatura\n"
        "• Comprobar ruidos anómalos\n"
        "• Inspeccionar tablero de control",
        parse_mode='HTML'
    )
    return ESPERANDO_ITEMS_CHECKLIST

async def recibir_items_checklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recibe los ítems del checklist"""
    items_texto = update.message.text.strip()
    items = [item.strip().lstrip('•-*').strip() for item in items_texto.split('\n') if item.strip()]
    
    if len(items) < 2:
        await update.message.reply_text(
            "⚠️ Debes ingresar al menos 2 ítems.\n"
            "Escribe los ítems nuevamente, uno por línea:"
        )
        return ESPERANDO_ITEMS_CHECKLIST
    
    nombre_maquina = context.user_data['nombre_maquina']
    checklists = ChecklistManager.cargar_checklists()
    checklists[nombre_maquina] = items
    ChecklistManager.guardar_checklists(checklists)
    
    await update.message.reply_text(
        f"✅ **Máquina agregada exitosamente**\n\n"
        f"🏭 **Nombre:** {nombre_maquina}\n"
        f"📋 **Ítems:** {len(items)}\n\n"
        f"**Lista de verificación:**\n" + 
        '\n'.join(f"• {item}" for item in items),
        parse_mode='HTML'
    )
    
    return ConversationHandler.END

async def comando_ver_maquinas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra las máquinas configuradas"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden ver las máquinas.")
        return
    
    checklists = ChecklistManager.cargar_checklists()
    if not checklists:
        await update.message.reply_text("📋 No hay máquinas configuradas aún.")
        return
    
    mensaje = "🏭 **Máquinas configuradas:**\n\n"
    for nombre, items in checklists.items():
        mensaje += f"• **{nombre}** ({len(items)} ítems)\n"
    
    await update.message.reply_text(mensaje, parse_mode='HTML')

async def comando_eliminar_maquina(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Permite eliminar una máquina"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden eliminar máquinas.")
        return
    
    checklists = ChecklistManager.cargar_checklists()
    if not checklists:
        await update.message.reply_text("📋 No hay máquinas para eliminar.")
        return
    
    botones = []
    for maquina in checklists:
        botones.append([InlineKeyboardButton(f"🗑️ {maquina}", callback_data=f"eliminar_{maquina}")])
    
    botones.append([InlineKeyboardButton("❌ Cancelar", callback_data="cancelar_eliminacion")])
    
    keyboard = InlineKeyboardMarkup(botones)
    await update.message.reply_text(
        "🗑️ **Selecciona la máquina a eliminar:**\n\n"
        "⚠️ *Esta acción no se puede deshacer*",
        reply_markup=keyboard,
        parse_mode='HTML'
    )

async def comando_checklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra las máquinas disponibles para checklist"""
    checklists = ChecklistManager.cargar_checklists()
    if not checklists:
        await update.message.reply_text(
            "📋 No hay máquinas disponibles.\n\n"
            "Los administradores pueden agregar máquinas con /nueva_maquina"
        )
        return
    
    botones = []
    for maquina in checklists:
        botones.append([InlineKeyboardButton(f"🔧 {maquina}", callback_data=f"checklist_{maquina}")])
    
    keyboard = InlineKeyboardMarkup(botones)
    await update.message.reply_text(
        "🔧 **Selecciona una máquina para realizar el checklist:**",
        reply_markup=keyboard,
        parse_mode='HTML'
    )

async def manejar_callback_checklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja los callbacks de selección de máquina y eliminación"""
    query = update.callback_query
    await query.answer()
    data = query.data
    
    if data.startswith("checklist_"):
        maquina = data.replace("checklist_", "")
        await iniciar_checklist(query, context, maquina)
    
    elif data.startswith("eliminar_"):
        maquina = data.replace("eliminar_", "")
        if ChecklistManager.eliminar_maquina(maquina):
            await query.edit_message_text(f"✅ Máquina '{maquina}' eliminada exitosamente.")
        else:
            await query.edit_message_text(f"❌ Error al eliminar la máquina '{maquina}'.")
    
    elif data == "cancelar_eliminacion":
        await query.edit_message_text("❌ Eliminación cancelada.")

async def iniciar_checklist(query, context, maquina):
    """Inicia el proceso de checklist para una máquina"""
    user_id = query.from_user.id
    operador = query.from_user.first_name
    
    checklists = ChecklistManager.cargar_checklists()
    if maquina not in checklists:
        await query.edit_message_text("❌ Máquina no encontrada.")
        return
    
    items = checklists[maquina]
    checklists_activos[user_id] = {
        "maquina": maquina,
        "operador": operador,
        "items": items,
        "item_actual": 0,
        "resultados": [],
        "inicio": datetime.now(),
        "observaciones_pendientes": False,
        "esperando_foto": False  # Nuevo estado para fotos
    }
    
    await mostrar_item_checklist(query, context, user_id)

async def mostrar_item_checklist(query, context, user_id):
    """Muestra el ítem actual del checklist"""
    if user_id not in checklists_activos:
        await query.edit_message_text("❌ Sesión de checklist no encontrada.")
        return
    
    activo = checklists_activos[user_id]
    idx = activo['item_actual']
    items = activo['items']
    
    if idx >= len(items):
        await finalizar_checklist(query, context, user_id)
        return
    
    item = items[idx]
    progreso = f"[{idx + 1}/{len(items)}]"
    
    botones = [
        [InlineKeyboardButton("✅ OK", callback_data="resp_OK")],
        [InlineKeyboardButton("⚠️ Revisar", callback_data="resp_REVISAR")],
        [InlineKeyboardButton("❌ Falla", callback_data="resp_FALLA")],
        [InlineKeyboardButton("🔄 Cancelar Checklist", callback_data="cancelar_checklist")]
    ]
    
    keyboard = InlineKeyboardMarkup(botones)
    mensaje = f"🔧 **{activo['maquina']}** {progreso}\n\n📋 **{item}**"
    
    await query.edit_message_text(mensaje, reply_markup=keyboard, parse_mode='HTML')

async def procesar_respuesta_checklist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Procesa la respuesta del usuario al ítem del checklist"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    if user_id not in checklists_activos:
        await query.edit_message_text("❌ Sesión de checklist no encontrada.")
        return
    
    data = query.data
    
    if data == "cancelar_checklist":
        del checklists_activos[user_id]
        await query.edit_message_text("🔄 Checklist cancelado.")
        return
    
    respuesta = data.replace("resp_", "")
    activo = checklists_activos[user_id]
    idx = activo['item_actual']
    item = activo['items'][idx]
    
    estados = {
        "OK": "✅ OK",
        "REVISAR": "⚠️ REVISAR", 
        "FALLA": "❌ FALLA"
    }
    
    estado = estados.get(respuesta, respuesta)
    
    # Si es REVISAR o FALLA, ofrecer opciones para documentar
    if respuesta in ["REVISAR", "FALLA"]:
        activo['respuesta_temporal'] = {
            "item": item,
            "estado": estado,
            "hora": datetime.now().strftime("%H:%M:%S"),
            "observaciones": "",
            "foto_ruta": ""
        }
        activo['observaciones_pendientes'] = True
        
        # Botones para documentar el problema
        botones = [
            [InlineKeyboardButton("📸 Enviar Foto", callback_data="solicitar_foto")],
            [InlineKeyboardButton("📝 Solo Observaciones", callback_data="solo_observaciones")],
            [InlineKeyboardButton("⏭️ Continuar Sin Documentar", callback_data="continuar_sin_doc")]
        ]
        
        keyboard = InlineKeyboardMarkup(botones)
        await query.edit_message_text(
            f"📋 **{item}**\n"
            f"Estado: {estado}\n\n"
            f"🔍 **¿Cómo quieres documentar este problema?**\n\n"
            f"📸 Puedes enviar una foto del problema\n"
            f"📝 O escribir observaciones de texto\n"
            f"⏭️ O continuar sin documentar",
            reply_markup=keyboard,
            parse_mode='HTML'
        )
        return
    
    # Si es OK, continuar directamente
    activo['resultados'].append({
        "item": item,
        "estado": estado,
        "observaciones": "",
        "foto_ruta": "",
        "hora": datetime.now().strftime("%H:%M:%S")
    })
    
    activo['item_actual'] += 1
    await mostrar_item_checklist(query, context, user_id)

async def manejar_documentacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja las opciones de documentación de problemas"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    if user_id not in checklists_activos:
        await query.edit_message_text("❌ Sesión de checklist no encontrada.")
        return
    
    activo = checklists_activos[user_id]
    data = query.data
    
    if data == "solicitar_foto":
        activo['esperando_foto'] = True
        await query.edit_message_text(
            f"📸 **Envía una foto del problema**\n\n"
            f"📋 Item: {activo['respuesta_temporal']['item']}\n"
            f"⚠️ Estado: {activo['respuesta_temporal']['estado']}\n\n"
            f"📷 Envía la foto como imagen (no como archivo)\n"
            f"📝 Después podrás agregar observaciones de texto\n\n"
            f"💡 Tip: Asegúrate de que la foto sea clara y muestre el problema",
            parse_mode='HTML'
        )
        
    elif data == "solo_observaciones":
        await query.edit_message_text(
            f"📝 **Escribe las observaciones**\n\n"
            f"📋 Item: {activo['respuesta_temporal']['item']}\n"
            f"⚠️ Estado: {activo['respuesta_temporal']['estado']}\n\n"
            f"Describe el problema encontrado:",
            parse_mode='HTML'
        )
        
    elif data == "continuar_sin_doc":
        # Guardar sin documentación
        resp_temp = activo['respuesta_temporal']
        activo['resultados'].append(resp_temp)
        activo['item_actual'] += 1
        activo['observaciones_pendientes'] = False
        del activo['respuesta_temporal']
        
        await query.edit_message_text("⏭️ Continuando sin documentar...")
        await mostrar_item_checklist_nuevo_mensaje(query, context, user_id)

# Agregar handler para manejar documentación
async def callback_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler unificado para callbacks"""
    query = update.callback_query
    data = query.data
    
    if data.startswith("checklist_") or data.startswith("eliminar_") or data == "cancelar_eliminacion":
        await manejar_callback_checklist(update, context)
    elif data.startswith("resp_"):
        await procesar_respuesta_checklist(update, context)
    elif data in ["solicitar_foto", "solo_observaciones", "continuar_sin_doc"]:
        await manejar_documentacion(update, context)

async def manejar_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja las fotos enviadas por el usuario"""
    user_id = update.message.from_user.id
    
    if (user_id not in checklists_activos or 
        not checklists_activos[user_id].get('esperando_foto')):
        # Foto enviada fuera de contexto
        await update.message.reply_text(
            "📸 Foto recibida, pero no hay un checklist activo que requiera fotos.\n"
            "Usa /checklist para iniciar una verificación."
        )
        return
    
    activo = checklists_activos[user_id]
    
    try:
        # Obtener el archivo de foto de mayor resolución
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        
        # Generar nombre único para la foto
        nombre_foto = FotoManager.generar_nombre_foto(
            activo['maquina'], 
            activo['respuesta_temporal']['item'],
            user_id
        )
        
        # Guardar la foto
        ruta_foto = await FotoManager.guardar_foto(file, nombre_foto)
        
        if ruta_foto:
            # Actualizar la respuesta temporal con la ruta de la foto
            activo['respuesta_temporal']['foto_ruta'] = FotoManager.obtener_ruta_relativa(ruta_foto)
            activo['esperando_foto'] = False
            
            await update.message.reply_text(
                f"✅ **Foto guardada exitosamente**\n\n"
                f"📋 Item: {activo['respuesta_temporal']['item']}\n"
                f"⚠️ Estado: {activo['respuesta_temporal']['estado']}\n"
                f"📸 Foto: {nombre_foto}\n\n"
                f"📝 Ahora escribe las observaciones sobre el problema encontrado:\n"
                f"(Puedes escribir 'sin observaciones' si no tienes más detalles)",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                "❌ Error al guardar la foto. Intenta nuevamente o continúa sin foto.\n"
                "Escribe 'continuar' para seguir sin foto."
            )
    
    except Exception as e:
        logger.error(f"Error procesando foto: {e}")
        await update.message.reply_text(
            "❌ Error al procesar la foto. Intenta nuevamente o continúa sin foto.\n"
            "Escribe 'continuar' para seguir sin foto."
        )

async def manejar_observaciones(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja las observaciones de texto enviadas por el usuario"""
    user_id = update.message.from_user.id
    
    if (user_id not in checklists_activos or 
        not checklists_activos[user_id].get('observaciones_pendientes')):
        return  # No es una observación para checklist
    
    activo = checklists_activos[user_id]
    observaciones = update.message.text.strip()
    
    # Verificar si el usuario quiere continuar sin observaciones
    if observaciones.lower() in ['continuar', 'sin observaciones', 'ninguna', 'no']:
        observaciones = ""
    
    # Guardar las observaciones
    activo['respuesta_temporal']['observaciones'] = observaciones
    
    # Finalizar este item y continuar
    activo['resultados'].append(activo['respuesta_temporal'])
    activo['item_actual'] += 1
    activo['observaciones_pendientes'] = False
    del activo['respuesta_temporal']
    
    # Mensaje de confirmación
    confirmacion = f"✅ **Documentación guardada**\n\n"
    if observaciones:
        confirmacion += f"📝 Observaciones: {observaciones[:100]}{'...' if len(observaciones) > 100 else ''}\n"
    if activo['resultados'][-1]['foto_ruta']:
        confirmacion += f"📸 Foto incluida\n"
    confirmacion += f"\n⏭️ Continuando con el siguiente item..."
    
    await update.message.reply_text(confirmacion, parse_mode='HTML')
    
    # Continuar con el siguiente item
    await mostrar_item_checklist_texto(update, context, user_id)

async def mostrar_item_checklist_texto(update, context, user_id):
    """Muestra el siguiente item del checklist en un mensaje de texto"""
    if user_id not in checklists_activos:
        await update.message.reply_text("❌ Sesión de checklist no encontrada.")
        return
    
    activo = checklists_activos[user_id]
    idx = activo['item_actual']
    items = activo['items']
    
    if idx >= len(items):
        await finalizar_checklist_texto(update, context, user_id)
        return
    
    item = items[idx]
    progreso = f"[{idx + 1}/{len(items)}]"
    
    botones = [
        [InlineKeyboardButton("✅ OK", callback_data="resp_OK")],
        [InlineKeyboardButton("⚠️ Revisar", callback_data="resp_REVISAR")],
        [InlineKeyboardButton("❌ Falla", callback_data="resp_FALLA")],
        [InlineKeyboardButton("🔄 Cancelar Checklist", callback_data="cancelar_checklist")]
    ]
    
    keyboard = InlineKeyboardMarkup(botones)
    mensaje = f"🔧 **{activo['maquina']}** {progreso}\n\n📋 **{item}**"
    
    await update.message.reply_text(mensaje, reply_markup=keyboard, parse_mode='HTML')

async def mostrar_item_checklist_nuevo_mensaje(query, context, user_id):
    """Versión para callback que crea un nuevo mensaje"""
    if user_id not in checklists_activos:
        return
    
    activo = checklists_activos[user_id]
    idx = activo['item_actual']
    items = activo['items']
    
    if idx >= len(items):
        await finalizar_checklist_callback(query, context, user_id)
        return
    
    item = items[idx]
    progreso = f"[{idx + 1}/{len(items)}]"
    
    botones = [
        [InlineKeyboardButton("✅ OK", callback_data="resp_OK")],
        [InlineKeyboardButton("⚠️ Revisar", callback_data="resp_REVISAR")],
        [InlineKeyboardButton("❌ Falla", callback_data="resp_FALLA")],
        [InlineKeyboardButton("🔄 Cancelar Checklist", callback_data="cancelar_checklist")]
    ]
    
    keyboard = InlineKeyboardMarkup(botones)
    mensaje = f"🔧 **{activo['maquina']}** {progreso}\n\n📋 **{item}**"
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=mensaje,
        reply_markup=keyboard,
        parse_mode='HTML'
    )

async def finalizar_checklist(query, context, user_id):
    """Finaliza el checklist y guarda los resultados"""
    activo = checklists_activos[user_id]
    fin = datetime.now()
    duracion = str(fin - activo['inicio']).split('.')[0]
    
    # Analizar resultados
    total_items = len(activo['resultados'])
    ok_count = sum(1 for r in activo['resultados'] if 'OK' in r['estado'])
    revisar_count = sum(1 for r in activo['resultados'] if 'REVISAR' in r['estado'])
    falla_count = sum(1 for r in activo['resultados'] if 'FALLA' in r['estado'])
    fotos_count = sum(1 for r in activo['resultados'] if r['foto_ruta'])
    
    # Determinar resultado final
    if falla_count > 0:
        resultado_final = f"❌ REQUIERE ATENCIÓN ({falla_count} fallas)"
    elif revisar_count > 0:
        resultado_final = f"⚠️ REVISAR ({revisar_count} items)"
    else:
        resultado_final = "✅ APROBADO"
    
    # Guardar cada resultado en Excel
    fecha = fin.strftime("%d/%m/%Y")
    for resultado in activo['resultados']:
        ExcelManager.guardar_resultado_checklist(
            fecha=fecha,
            hora=resultado['hora'],
            operador=activo['operador'],
            maquina=activo['maquina'],
            item=resultado['item'],
            estado=resultado['estado'],
            observaciones=resultado['observaciones'],
            foto_ruta=resultado['foto_ruta'],
            resultado_final=resultado_final,
            duracion=duracion
        )
    
    # Mensaje de resumen
    resumen = (
        f"✅ **Checklist completado**\n\n"
        f"🏭 **Máquina:** {activo['maquina']}\n"
        f"👤 **Operador:** {activo['operador']}\n"
        f"⏱️ **Duración:** {duracion}\n"
        f"📊 **Resultado:** {resultado_final}\n\n"
        f"📋 **Resumen:**\n"
        f"• ✅ OK: {ok_count}\n"
        f"• ⚠️ Revisar: {revisar_count}\n"
        f"• ❌ Fallas: {falla_count}\n"
        f"• 📸 Fotos: {fotos_count}\n\n"
        f"💾 Datos guardados en el registro de mantenimiento."
    )
    
    await query.edit_message_text(resumen, parse_mode='HTML')
    del checklists_activos[user_id]

async def finalizar_checklist_texto(update, context, user_id):
    """Versión para mensajes de texto"""
    activo = checklists_activos[user_id]
    fin = datetime.now()
    duracion = str(fin - activo['inicio']).split('.')[0]
    
    # Analizar resultados
    total_items = len(activo['resultados'])
    ok_count = sum(1 for r in activo['resultados'] if 'OK' in r['estado'])
    revisar_count = sum(1 for r in activo['resultados'] if 'REVISAR' in r['estado'])
    falla_count = sum(1 for r in activo['resultados'] if 'FALLA' in r['estado'])
    fotos_count = sum(1 for r in activo['resultados'] if r['foto_ruta'])
    
    # Determinar resultado final
    if falla_count > 0:
        resultado_final = f"❌ REQUIERE ATENCIÓN ({falla_count} fallas)"
    elif revisar_count > 0:
        resultado_final = f"⚠️ REVISAR ({revisar_count} items)"
    else:
        resultado_final = "✅ APROBADO"
    
    # Guardar cada resultado en Excel
    fecha = fin.strftime("%d/%m/%Y")
    for resultado in activo['resultados']:
        ExcelManager.guardar_resultado_checklist(
            fecha=fecha,
            hora=resultado['hora'],
            operador=activo['operador'],
            maquina=activo['maquina'],
            item=resultado['item'],
            estado=resultado['estado'],
            observaciones=resultado['observaciones'],
            foto_ruta=resultado['foto_ruta'],
            resultado_final=resultado_final,
            duracion=duracion
        )
    
    # Mensaje de resumen
    resumen = (
        f"✅ **Checklist completado**\n\n"
        f"🏭 **Máquina:** {activo['maquina']}\n"
        f"👤 **Operador:** {activo['operador']}\n"
        f"⏱️ **Duración:** {duracion}\n"
        f"📊 **Resultado:** {resultado_final}\n\n"
        f"📋 **Resumen:**\n"
        f"• ✅ OK: {ok_count}\n"
        f"• ⚠️ Revisar: {revisar_count}\n"
        f"• ❌ Fallas: {falla_count}\n"
        f"• 📸 Fotos: {fotos_count}\n\n"
        f"💾 Datos guardados en el registro de mantenimiento."
    )
    
    await update.message.reply_text(resumen, parse_mode='HTML')
    del checklists_activos[user_id]

async def finalizar_checklist_callback(query, context, user_id):
    """Versión para callbacks"""
    activo = checklists_activos[user_id]
    fin = datetime.now()
    duracion = str(fin - activo['inicio']).split('.')[0]
    
    # Analizar resultados
    total_items = len(activo['resultados'])
    ok_count = sum(1 for r in activo['resultados'] if 'OK' in r['estado'])
    revisar_count = sum(1 for r in activo['resultados'] if 'REVISAR' in r['estado'])
    falla_count = sum(1 for r in activo['resultados'] if 'FALLA' in r['estado'])
    fotos_count = sum(1 for r in activo['resultados'] if r['foto_ruta'])
    
    # Determinar resultado final
    if falla_count > 0:
        resultado_final = f"❌ REQUIERE ATENCIÓN ({falla_count} fallas)"
    elif revisar_count > 0:
        resultado_final = f"⚠️ REVISAR ({revisar_count} items)"
    else:
        resultado_final = "✅ APROBADO"
    
    # Guardar cada resultado en Excel
    fecha = fin.strftime("%d/%m/%Y")
    for resultado in activo['resultados']:
        ExcelManager.guardar_resultado_checklist(
            fecha=fecha,
            hora=resultado['hora'],
            operador=activo['operador'],
            maquina=activo['maquina'],
            item=resultado['item'],
            estado=resultado['estado'],
            observaciones=resultado['observaciones'],
            foto_ruta=resultado['foto_ruta'],
            resultado_final=resultado_final,
            duracion=duracion
        )
    
    # Mensaje de resumen
    resumen = (
        f"✅ **Checklist completado**\n\n"
        f"🏭 **Máquina:** {activo['maquina']}\n"
        f"👤 **Operador:** {activo['operador']}\n"
        f"⏱️ **Duración:** {duracion}\n"
        f"📊 **Resultado:** {resultado_final}\n\n"
        f"📋 **Resumen:**\n"
        f"• ✅ OK: {ok_count}\n"
        f"• ⚠️ Revisar: {revisar_count}\n"
        f"• ❌ Fallas: {falla_count}\n"
        f"• 📸 Fotos: {fotos_count}\n\n"
        f"💾 Datos guardados en el registro de mantenimiento."
    )
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=resumen,
        parse_mode='HTML'
    )
    del checklists_activos[user_id]

async def comando_descargar_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Permite descargar el archivo Excel"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden descargar el registro.")
        return
    
    if not os.path.exists(ARCHIVO_EXCEL):
        await update.message.reply_text("📋 No hay datos de mantenimiento registrados aún.")
        return
    
    try:
        await update.message.reply_document(
            document=open(ARCHIVO_EXCEL, 'rb'),
            filename=f"mantenimiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            caption="📊 Registro de mantenimiento actualizado"
        )
        logger.info(f"Archivo Excel enviado a {update.message.from_user.first_name}")
    except Exception as e:
        logger.error(f"Error enviando Excel: {e}")
        await update.message.reply_text("❌ Error al enviar el archivo.")

async def comando_estadisticas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra estadísticas del mantenimiento"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden ver estadísticas.")
        return
    
    if not os.path.exists(ARCHIVO_EXCEL):
        await update.message.reply_text("📋 No hay datos para mostrar estadísticas.")
        return
    
    try:
        # Leer Excel y generar estadísticas básicas
        libro = openpyxl.load_workbook(ARCHIVO_EXCEL)
        hoja = libro.active
        
        total_registros = hoja.max_row - 4  # Restar encabezados
        if total_registros <= 0:
            await update.message.reply_text("📋 No hay registros de mantenimiento.")
            return
        
        # Contar estados (columna F - índice 6)
        ok_count = 0
        revisar_count = 0
        falla_count = 0
        fotos_count = 0
        
        for fila in range(5, hoja.max_row + 1):
            estado = hoja.cell(row=fila, column=6).value or ""
            foto = hoja.cell(row=fila, column=8).value or ""
            
            if "OK" in str(estado):
                ok_count += 1
            elif "REVISAR" in str(estado):
                revisar_count += 1
            elif "FALLA" in str(estado):
                falla_count += 1
            
            if foto:
                fotos_count += 1
        
        porcentaje_ok = (ok_count / total_registros * 100) if total_registros > 0 else 0
        
        mensaje = (
            f"📊 **Estadísticas de Mantenimiento**\n\n"
            f"📋 **Total de verificaciones:** {total_registros}\n\n"
            f"✅ **OK:** {ok_count} ({porcentaje_ok:.1f}%)\n"
            f"⚠️ **Revisar:** {revisar_count}\n"
            f"❌ **Fallas:** {falla_count}\n"
            f"📸 **Con fotos:** {fotos_count}\n\n"
            f"📈 **Eficiencia:** {porcentaje_ok:.1f}% de items OK"
        )
        
        await update.message.reply_text(mensaje, parse_mode='HTML')
        
    except Exception as e:
        logger.error(f"Error generando estadísticas: {e}")
        await update.message.reply_text("❌ Error al generar estadísticas.")

async def comando_agregar_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inicia el proceso para agregar un nuevo administrador"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden agregar otros administradores.")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "👤 **Agregar nuevo administrador**\n\n"
        "Escribe el ID del usuario que quieres hacer administrador:\n\n"
        "💡 *Tip: El usuario puede ver su ID con el comando /mi_id*",
        parse_mode='HTML'
    )
    return ESPERANDO_NUEVO_ADMIN

async def recibir_nuevo_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recibe el ID del nuevo administrador"""
    try:
        nuevo_admin_id = int(update.message.text.strip())
    except ValueError:
        await update.message.reply_text(
            "❌ ID inválido. Debe ser un número.\n"
            "Escribe el ID nuevamente:"
        )
        return ESPERANDO_NUEVO_ADMIN
    
    config = ConfigManager.cargar_config()
    admin_ids = config.get("admin_ids", ADMIN_IDS)
    
    if nuevo_admin_id in admin_ids:
        await update.message.reply_text(f"⚠️ El usuario {nuevo_admin_id} ya es administrador.")
        return ConversationHandler.END
    
    admin_ids.append(nuevo_admin_id)
    config["admin_ids"] = admin_ids
    ConfigManager.guardar_config(config)
    
    await update.message.reply_text(
        f"✅ **Administrador agregado**\n\n"
        f"👤 Nuevo admin ID: `{nuevo_admin_id}`\n"
        f"📋 Total administradores: {len(admin_ids)}",
        parse_mode='HTML'
    )
    
    return ConversationHandler.END

async def comando_ver_fotos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Muestra las fotos más recientes"""
    if not es_admin(update.message.from_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden ver las fotos.")
        return
    
    if not os.path.exists(CARPETA_FOTOS):
        await update.message.reply_text("📸 No hay fotos guardadas.")
        return
    
    fotos = os.listdir(CARPETA_FOTOS)
    if not fotos:
        await update.message.reply_text("📸 No hay fotos guardadas.")
        return
    
    # Ordenar por fecha (más recientes primero)
    fotos.sort(reverse=True)
    fotos_recientes = fotos[:5]  # Mostrar solo las 5 más recientes
    
    mensaje = f"📸 **Fotos recientes** ({len(fotos)} total)\n\n"
    for foto in fotos_recientes:
        # Extraer información del nombre del archivo
        partes = foto.replace('.jpg', '').split('_')
        if len(partes) >= 4:
            fecha_hora = partes[0] + '_' + partes[1]
            try:
                dt = datetime.strptime(fecha_hora, '%Y%m%d_%H%M%S')
                fecha_formateada = dt.strftime('%d/%m/%Y %H:%M')
                maquina = partes[2].replace('_', ' ')
                mensaje += f"• {fecha_formateada} - {maquina}\n"
            except:
                mensaje += f"• {foto}\n"
        else:
            mensaje += f"• {foto}\n"
    
    if len(fotos) > 5:
        mensaje += f"\n📁 Y {len(fotos) - 5} fotos más..."
    
    await update.message.reply_text(mensaje, parse_mode='HTML')

async def manejar_mensaje_texto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja mensajes de texto que pueden ser observaciones"""
    await manejar_observaciones(update, context)

# ============================================
# API FLASK INTEGRADA
# ============================================

app_flask = Flask(__name__)
CORS(app_flask)

@app_flask.route('/api/bot/status', methods=['GET'])
def get_bot_status():
    """Estado del bot"""
    uptime = None
    if bot_stats["inicio_bot"]:
        uptime = str(datetime.now() - bot_stats["inicio_bot"]).split('.')[0]
    
    return jsonify({
        "is_running": True,
        "status": "running",
        "start_time": bot_stats["inicio_bot"].isoformat() if bot_stats["inicio_bot"] else None,
        "uptime": uptime,
        "stats": bot_stats
    })

@app_flask.route('/api/telegram/stats', methods=['GET'])
def get_telegram_stats():
    """Estadísticas del sistema"""
    stats = {
        "total_maquinas": 0,
        "total_checklists": 0,
        "total_fotos": 0,
        "ultimo_checklist": None,
        "maquinas_detalles": []
    }
    
    # Máquinas configuradas
    if os.path.exists(ARCHIVO_CHECKLISTS):
        try:
            with open(ARCHIVO_CHECKLISTS, 'r', encoding='utf-8') as f:
                maquinas = json.load(f)
                stats["total_maquinas"] = len(maquinas)
                stats["maquinas_detalles"] = [
                    {"nombre": nombre, "items": len(items)} 
                    for nombre, items in maquinas.items()
                ]
        except:
            pass
    
    # Fotos guardadas
    if os.path.exists(CARPETA_FOTOS):
        fotos = os.listdir(CARPETA_FOTOS)
        stats["total_fotos"] = len([f for f in fotos if f.endswith('.jpg')])
    
    # Estadísticas de Excel
    if os.path.exists(ARCHIVO_EXCEL):
        try:
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL)
            hoja = libro.active
            stats["total_checklists"] = max(0, hoja.max_row - 4)
            
            if stats["total_checklists"] > 0:
                ultima_fila = hoja.max_row
                stats["ultimo_checklist"] = {
                    "fecha": str(hoja.cell(row=ultima_fila, column=1).value or ""),
                    "operador": str(hoja.cell(row=ultima_fila, column=3).value or ""),
                    "maquina": str(hoja.cell(row=ultima_fila, column=4).value or ""),
                    "resultado": str(hoja.cell(row=ultima_fila, column=9).value or "")
                }
        except:
            pass
    
    return jsonify(stats)

@app_flask.route('/api/sync/export-to-telegram', methods=['POST'])
def export_to_telegram():
    """Exporta datos del HTML al sistema"""
    html_data = request.json
    
    try:
        telegram_checklists = {}
        
        for equipo in html_data.get('equipos', []):
            items_basicos = [
                f"Verificar estado general de {equipo['nombre']}",
                f"Comprobar funcionamiento de {equipo['tipo']}",
                "Revisar niveles de lubricación",
                "Inspeccionar conexiones eléctricas",
                "Verificar temperatura de operación",
                "Comprobar ruidos anómalos",
                "Inspeccionar elementos de seguridad"
            ]
            telegram_checklists[equipo['nombre']] = items_basicos
        
        if telegram_checklists:
            with open(ARCHIVO_CHECKLISTS, 'w', encoding='utf-8') as f:
                json.dump(telegram_checklists, f, ensure_ascii=False, indent=2)
            
            actualizar_stats()
            
            return jsonify({
                "success": True,
                "message": f"Exportadas {len(telegram_checklists)} máquinas al bot"
            })
        
        return jsonify({"success": False, "message": "No hay máquinas para exportar"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error: {str(e)}"})

@app_flask.route('/api/files/excel', methods=['GET'])
def download_excel():
    """Descargar Excel"""
    if os.path.exists(ARCHIVO_EXCEL):
        from flask import send_file
        return send_file(
            ARCHIVO_EXCEL,
            as_attachment=True,
            download_name=f"mantenimiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    else:
        return jsonify({"error": "Archivo Excel no encontrado"}), 404

def actualizar_stats():
    """Actualiza las estadísticas globales"""
    try:
        if os.path.exists(ARCHIVO_CHECKLISTS):
            with open(ARCHIVO_CHECKLISTS, 'r', encoding='utf-8') as f:
                checklists = json.load(f)
                bot_stats["maquinas_configuradas"] = len(checklists)
        
        if os.path.exists(CARPETA_FOTOS):
            fotos = os.listdir(CARPETA_FOTOS)
            bot_stats["fotos_guardadas"] = len([f for f in fotos if f.endswith('.jpg')])
        
        if os.path.exists(ARCHIVO_EXCEL):
            libro = openpyxl.load_workbook(ARCHIVO_EXCEL)
            hoja = libro.active
            bot_stats["checklists_completados"] = max(0, hoja.max_row - 4)
    except:
        pass

def ejecutar_flask():
    """Ejecuta Flask en un hilo separado"""
    app_flask.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
@app.route('/bot/start', methods=['POST'])

def start_chatbot():
    try:
        # Aquí es donde inicias el bot
        application = Application.builder().token(TOKEN).build()
        application.run_polling()
        
        return jsonify({'status': 'success', 'message': 'Chatbot iniciado'}), 200
    except Exception as e:
        print(f"Error al iniciar el chatbot: {e}")
        return jsonify({'status': 'error', 'message': 'No se pudo iniciar el chatbot'}), 500

@app.route('/bot/stop', methods=['POST'])
def stop_chatbot():
    try:
        # Código para detener el bot si es necesario
        # Aquí puedes definir cómo detener el bot, por ejemplo, interrumpiendo el polling.
        # Esto generalmente no se recomienda, ya que el polling está bloqueando la ejecución.
        return jsonify({'status': 'success', 'message': 'Chatbot detenido'}), 200
    except Exception as e:
        print(f"Error al detener el chatbot: {e}")
        return jsonify({'status': 'error', 'message': 'No se pudo detener el chatbot'}), 500

def main():
    """Función principal"""
    # Inicializar estadísticas
    bot_stats["inicio_bot"] = datetime.now()
    actualizar_stats()
    
    # Iniciar Flask en un hilo separado
    flask_thread = threading.Thread(target=ejecutar_flask, daemon=True)
    flask_thread.start()
    
    print("🚀 Bot de mantenimiento iniciado")
    print("🌐 API web disponible en: http://localhost:5000")
    print("📱 Bot de Telegram activado")
    
    # Crear la aplicación
    application = ApplicationBuilder().token(TOKEN).build()
    
    # Handlers de comandos
    application.add_handler(CommandHandler("start", comando_start))
    application.add_handler(CommandHandler("mi_id", comando_mi_id))
    application.add_handler(CommandHandler("checklist", comando_checklist))
    application.add_handler(CommandHandler("ver_maquinas", comando_ver_maquinas))
    application.add_handler(CommandHandler("eliminar_maquina", comando_eliminar_maquina))
    application.add_handler(CommandHandler("descargar_excel", comando_descargar_excel))
    application.add_handler(CommandHandler("estadisticas", comando_estadisticas))
    application.add_handler(CommandHandler("ver_fotos", comando_ver_fotos))
    
    # Handler de conversación para nueva máquina
    conv_nueva_maquina = ConversationHandler(
        entry_points=[CommandHandler("nueva_maquina", comando_nueva_maquina)],
        states={
            ESPERANDO_NOMBRE_MAQUINA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nombre_maquina)],
            ESPERANDO_ITEMS_CHECKLIST: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_items_checklist)],
        },
        fallbacks=[CommandHandler("cancelar", comando_start)]
    )
    application.add_handler(conv_nueva_maquina)
    
    # Handler de conversación para agregar admin
    conv_agregar_admin = ConversationHandler(
        entry_points=[CommandHandler("agregar_admin", comando_agregar_admin)],
        states={
            ESPERANDO_NUEVO_ADMIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nuevo_admin)],
        },
        fallbacks=[CommandHandler("cancelar", comando_start)]
    )
    application.add_handler(conv_agregar_admin)
    
    # Handlers para callbacks y fotos
    application.add_handler(CallbackQueryHandler(callback_query_handler))
    application.add_handler(MessageHandler(filters.PHOTO, manejar_foto))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_mensaje_texto))
    
    # Iniciar el bot
    logger.info("🤖 Bot de mantenimiento iniciado")
    application.run_polling()

if __name__ == '__main__':
    main()